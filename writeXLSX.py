import openpyxl
def write07Excel(path, data, title):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = title
    sheet.cell(row=1, column=1, value="书名")
    sheet.cell(row=1, column=2, value="作者名")
    sheet.cell(row=1, column=3, value="出版时间")
    sheet.cell(row=1, column=4, value="售价")
    sheet.cell(row=1, column=5, value="出版社")
    sheet.cell(row=1, column=6, value="丛书名")
    sheet.cell(row=1, column=7, value="内容简介")
    sheet.cell(row=1, column=8, value="ISBN")

    for i in range(0, len(data)):
        sheet.cell(row=i + 2, column=1, value=data[i].name)
        sheet.cell(row=i + 2, column=2, value=data[i].author)
        sheet.cell(row=i + 2, column=3, value=data[i].publication_date)
        sheet.cell(row=i + 2, column=4, value=data[i].price)
        sheet.cell(row=i + 2, column=5, value=data[i].publish_company)
        sheet.cell(row=i + 2, column=6, value=data[i].books_name)
        sheet.cell(row=i + 2, column=7, value=data[i].content)
        sheet.cell(row=i + 2, column=8, value=data[i].ISBN)

    wb.save(path)
    print("写入数据成功！")

# write07Excel("")